import re
import sys
import os

try:
    from PyPDF2 import PdfReader
except ImportError:
    os.system("pip install PyPDF2")
    from PyPDF2 import PdfReader

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


DESIGNATION_SHORT = {
    'PRIMARY SCHOOL HEAD TEACHER': 'PSHT',
    'PRIMARY SCHOOL HEAD TEACH': 'PSHT',
    'PRIMARY SCHOOL HEADTEACHER': 'PSHT',
    'SENIOR PRIMARY SCHOOL TEACHER': 'SPST',
    'SENIOR PRIMARY SCHOOL TEACH': 'SPST',
    'SR PRIMARY SCHOOL TEACHER': 'SPST',
    'PRIMARY SCHOOL TEACHER': 'PST',
    'PRIMARY SCHOOL TEACH': 'PST',
    'SECONDARY SCHOOL TEACHER': 'SST',
    'SECONDARY SCHOOL TEACH': 'SST',
    'SENIOR CERTIFIED TEACHER': 'SCT',
    'SENIOR CERTIFIED TEACH': 'SCT',
    'CERTIFIED TEACHER': 'CT',
    'CERTIFIED TEACH': 'CT',
    'JUNIOR CERTIFIED TEACHER': 'JCT',
    'JUNIOR CERTIFIED TEACH': 'JCT',
    'PHYSICAL TRAINING INSTRUCTOR': 'PTI',
    'PHYSICAL TRAINING INSTRUCT': 'PTI',
    'DRAWING MASTER': 'DM',
    'QARI': 'QARI',
    'ARABIC TEACHER': 'AT',
    'ARABIC TEACH': 'AT',
    'HEAD MASTER': 'HM',
    'HEAD MISTRESS': 'HM',
    'HEADMASTER': 'HM',
    'HEADMISTRESS': 'HM',
    'PRINCIPAL': 'PRINCIPAL',
    'VICE PRINCIPAL': 'VP',
    'LECTURER': 'LECTURER',
    'SUBJECT SPECIALIST': 'SS',
    'SUPERINTENDENT': 'SUPERINTENDENT',
    'ASSISTANT SUB DIVISIONAL EDUCATION': 'ASDEO',
    'ASSISTANT SUB DIVISIONAL': 'ASDEO',
    'ASSISTANT SUB DIVISION': 'ASDEO',
    'SUB DIVISIONAL EDUCATION': 'SDEO',
    'SUB DIVISIONAL EDUCAT': 'SDEO',
    'DISTRICT EDUCATION OFFICER': 'DEO',
    'DISTRICT EDUCATION': 'DEO',
    'ASSISTANT': 'ASSISTANT',
    'COMPUTER OPERATOR': 'CO',
    'JUNIOR CLERK': 'JC',
    'SENIOR CLERK': 'SC',
    'STENOGRAPHER': 'STENO',
    'DRIVER': 'DRIVER',
    'NAIB QASID': 'NQ',
    'CHOWKIDAR': 'CHOWKIDAR',
    'SWEEPER': 'SWEEPER',
    'PEON': 'PEON',
    'MALI': 'MALI',
    'LAB ATTENDANT': 'LA',
    'LABORATORY ATTENDANT': 'LA',
    'LAB ASSISTANT': 'LA',
    'LIBRARY ASSISTANT': 'LIB ASST',
    'LIBRARIAN': 'LIBRARIAN',
}

TEACHING_DESIGNATIONS = [
    'PRIMARY SCHOOL HEAD TEACH',
    'SENIOR PRIMARY SCHOOL TEACH',
    'PRIMARY SCHOOL TEACH',
    'SECONDARY SCHOOL TEACH',
    'SENIOR CERTIFIED TEACH',
    'CERTIFIED TEACH',
    'JUNIOR CERTIFIED TEACH',
    'PHYSICAL TRAINING',
    'DRAWING MASTER',
    'QARI',
    'ARABIC TEACH',
    'HEAD MASTER', 'HEAD MISTRESS',
    'HEADMASTER', 'HEADMISTRESS',
    'PRINCIPAL',
    'VICE PRINCIPAL',
    'LECTURER',
    'SUBJECT SPECIALIST',
    'TEACHING ALLOWANCE',
    'TEACHER', 'TEACH',
    'PST', 'SST', 'CT', 'SCT', 'SPST', 'PSHT',
    'AT', 'DM', 'PTI', 'HM', 'SS',
]


# Mapping of SAP wage codes to internal allowance fields (aligned with PAY_ALLOWANCE_CODES)
PAY_ALLOWANCE_MAP = {
    '0001': 'basic_pay',
    '0046': 'p_pay',
    '1001': 'hra',
    '1210': 'ca',
    '1300': 'ma',
    '1505': 'charge_allow',
    '1551': 'spl_allow_disable',
    '1923': 'uaa',
    '2148': 'adhoc_2013',
    '2199': 'adhoc_10pct',
    '2311': 'dress_allow',
    '2312': 'wa',
    '2313': 'integrated_allow',
    '2316': 'teaching_allow',
    '2320': 'computer_allow',
    '2321': 'mphil_allow',
    '2322': 'entertainment_allow',
    '2323': 'science_teaching_allow',
    '2324': 'weather_allow',
    '2325': 'special_allow_non_teaching',
    '2341': 'dra_2022kp',
    '2347': 'adhoc_2022_ps17',
    '2348': 'adhoc_2015',
    '2349': 'adhoc_2016',
    '2350': 'adhoc_2022',
    '2378': 'adhoc_2023_35',
    '2393': 'adhoc_2024_25',
    '2419': 'adhoc_2025_10',
    '2431': 'dra_2025_15',
}

# Mapping of SAP deduction wage codes to internal deduction fields (aligned with DEDUCTION_CODES)
DEDUCTION_MAP = {
    '3004': 'gpf_sub',
    '3012': 'gpf_sub',
    '3013': 'gpf_sub',
    '3014': 'gpf_sub',
    '3015': 'gpf_sub',
    '3501': 'bf',
    '3609': 'income_tax_ded',
    '3914': 'edu_rop',
    '3990': 'eef',
    '4004': 'rb_death',
}


def get_short_designation(full_desig):
    if not full_desig:
        return ""
    upper = full_desig.upper().strip()
    if upper in DESIGNATION_SHORT:
        return DESIGNATION_SHORT[upper]
    best_match = ""
    best_short = upper
    for long_name, short_name in sorted(
        DESIGNATION_SHORT.items(), key=lambda x: len(x[0]), reverse=True
    ):
        if long_name in upper:
            if len(long_name) > len(best_match):
                best_match = long_name
                best_short = short_name
    return best_short


def get_staff_type(designation):
    if not designation:
        return ""
    upper = designation.upper().strip()
    for td in TEACHING_DESIGNATIONS:
        if td in upper:
            return "Teaching"
    return "Non-Teaching"


def format_cnic(cnic):
    if not cnic:
        return ""
    cnic = re.sub(r'[^0-9]', '', str(cnic))
    if len(cnic) == 13:
        return f"{cnic[0:5]}-{cnic[5:12]}-{cnic[12]}"
    return cnic


def format_date(date_str):
    if not date_str:
        return ""
    date_str = date_str.strip().replace('.', '/')
    parts = date_str.split('/')
    if len(parts) == 3:
        return f"{parts[0].zfill(2)}/{parts[1].zfill(2)}/{parts[2]}"
    return date_str


def clean_amount(text):
    if not text:
        return ""
    if isinstance(text, (int, float)):
        return abs(text)
    text = str(text).strip().replace(",", "").replace(" ", "")
    match = re.search(r'-?([\d]+\.?\d*)', text)
    if match:
        return float(match.group(1))
    return ""


def extract_section(text, start_marker, end_marker):
    start = text.find(start_marker)
    if start == -1:
        return ""
    end = text.find(end_marker, start + len(start_marker))
    if end == -1:
        return text[start:]
    return text[start:end]


def parse_pay_line(line):
    """Parse one line from pay/deduction table.
    
    Real PDF lines look like:
    '0001 Basic Pay    67,480.00 1001 House Rent Allowance 45%     3,524.00 '
    '1210 Convey Allowance  2005     2,856.00 1300 Medical Allowance     1,500.00 '
    '2148 15% Adhoc Relief All-2013       675.00 2199 Adhoc Relief Allow @10%       454.00 '
    '2431 Dispar. Red. All-15%-2025     3,588.00         0.00 '
    '3015 GPF Subscription     -4,290.00 3501 Benevolent Fund     -1,200.00'
    
    Rules for a valid wage code:
    - 4 digits
    - Must be at position 0 of line, OR right after previous amount+spaces
    - NOT a year embedded in description (like 2005, 2013, 2021, 2025)
    """
    results = {}
    
    # Step 1: Find ALL amounts with positions
    amounts = []
    for m in re.finditer(r'-?\s*[\d,]+\.\d{2}', line):
        raw = m.group().replace(' ', '').replace(',', '')
        try:
            val = float(raw)
        except ValueError:
            continue
        amounts.append({
            'start': m.start(),
            'end': m.end(),
            'val': abs(val),
            'raw': m.group()
        })
    
    if not amounts:
        return results
    
    # Step 2: Find valid wage code positions
    # A wage code is at line start OR immediately after whitespace
    # following a previous amount
    valid_code_starts = {0}  # position 0 is always valid
    for a in amounts:
        # After this amount ends, the next code can start
        # (with possible whitespace in between)
        pos = a['end']
        while pos < len(line) and line[pos] == ' ':
            pos += 1
        valid_code_starts.add(pos)
    
    # Step 3: Find all 4-digit sequences and check if at valid position
    codes_found = []
    for m in re.finditer(r'(\d{4})', line):
        code = m.group(1)
        pos = m.start()
        if pos in valid_code_starts:
            codes_found.append({
                'code': code,
                'start': pos,
                'end': m.end()
            })
    
    # Step 4: Pair each code with its amount
    # Each code's amount is the first amount that appears after it
    # but before the next code
    for i, cf in enumerate(codes_found):
        # Find boundary: next code start or end of line
        if i + 1 < len(codes_found):
            boundary = codes_found[i + 1]['start']
        else:
            boundary = len(line)
        
        # Find first amount between code end and boundary
        for a in amounts:
            if a['start'] >= cf['end'] and a['start'] < boundary:
                if a['val'] > 0:
                    results[cf['code']] = a['val']
                break
    
    return results


def parse_section_wages(section_text):
    wage = {}
    lines = section_text.split('\n')
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.startswith('Wage type') or line.startswith('Loan'):
            continue
        parsed = parse_pay_line(line)
        wage.update(parsed)
    return wage


def extract_employee_data(page_text):
    data = {}
    text = page_text

    pay_section = extract_section(
        text,
        'Wage type Amount Wage type Amount',
        'Deductions - General'
    )
    deduct_section = extract_section(
        text,
        'Deductions - General',
        'Deductions - Loans'
    )
    loan_section = extract_section(
        text,
        'Deductions - Loans and Advances',
        'Deductions - Income Tax'
    )

    pay_wages = parse_section_wages(pay_section)
    deduct_wages = parse_section_wages(deduct_section)

    # ===== name, father_name, gender =====
    name_match = re.search(
        r'Personal\s+Information\s+of\s+(Mr|Mrs|Ms|Miss)\s+'
        r'(.+?)\s+d/w/s\s+of\s+(.+)',
        text, re.IGNORECASE
    )
    if name_match:
        title = name_match.group(1).strip().lower()
        data['name'] = name_match.group(2).strip()
        father = name_match.group(3).strip().split('\n')[0].strip()
        father = re.sub(
            r'\s*Personnel.*', '', father, flags=re.IGNORECASE
        )
        data['father_name'] = father
        if title in ['mr']:
            data['gender'] = 'Male'
        elif title in ['mrs', 'ms', 'miss']:
            data['gender'] = 'Female'
        else:
            data['gender'] = ''
    else:
        data['name'] = ""
        data['father_name'] = ""
        data['gender'] = ""

    # ===== designation =====
    desig_match = re.search(
        r'Designation:\s*(.+?)\s+\d{8}-', text, re.IGNORECASE
    )
    raw_desig = desig_match.group(1).strip() if desig_match else ""
    data['designation'] = get_short_designation(raw_desig)
    data['designation_full'] = raw_desig

    # ===== bps =====
    bps_match = re.search(r'BPS:\s*(\d+)', text)
    data['bps'] = int(bps_match.group(1)) if bps_match else ""

    data['school_full_name'] = ""

    office_match = re.search(r'\d{8}-(.+?)(?:\n)', text)
    data['office_name'] = (
        office_match.group(1).strip() if office_match else ""
    )

    data['staff_type'] = get_staff_type(raw_desig)

    status_match = re.search(
        r'Employment\s+Category:\s*(.+?)(?:\n)', text, re.IGNORECASE
    )
    if status_match:
        raw = status_match.group(1).strip().lower()
        data['status'] = "Active" if 'active' in raw else "Inactive"
        if 'permanent' in raw:
            data['employment_category'] = "Permanent"
        elif 'temporary' in raw:
            data['employment_category'] = "Temporary"
        elif 'contract' in raw:
            data['employment_category'] = "Contract"
        else:
            data['employment_category'] = (
                status_match.group(1).strip()
            )
    else:
        data['status'] = ""
        data['employment_category'] = ""

    cnic_match = re.search(r'CNIC:\s*(\d{13})', text)
    data['cnic_no'] = (
        format_cnic(cnic_match.group(1)) if cnic_match else ""
    )

    pn_match = re.search(r'Personnel\s+Number:\s*(\d+)', text)
    data['personal_no'] = pn_match.group(1) if pn_match else ""

    data['mobile_no'] = ""

    dob_match = re.search(r'Date\s+of\s+Birth:\s*([\d.]+)', text)
    data['dob'] = (
        format_date(dob_match.group(1)) if dob_match else ""
    )

    data['nationality'] = "Pakistani"

    addr_match = re.search(
        r'Permanent\s+Address:\s*(.*?)(?:\n)', text
    )
    data['address'] = (
        addr_match.group(1).strip() if addr_match else ""
    )

    city_match = re.search(r'City:\s*(\w+)', text)
    data['district'] = (
        city_match.group(1).strip() if city_match else ""
    )

    data['tehsil'] = ""

    ddo_match = re.search(r'DDO\s+Code:\s*(\S+)', text)
    data['ddo_code'] = (
        ddo_match.group(1).strip().rstrip('-') if ddo_match else ""
    )

    acc_match = re.search(r'Account\s+Number:\s*(\S+)', text)
    data['bank_ac_no'] = acc_match.group(1) if acc_match else ""

    bank_match = re.search(
        r'Bank\s+Details:\s*(.+?)(?:\nLeaves)', text, re.DOTALL
    )
    if bank_match:
        bank_full = ' '.join(bank_match.group(1).split())
        parts = bank_full.split(',')
        data['bank_name'] = parts[0].strip() if parts else ""
        bc_match = re.search(r'(\d{6})', bank_full)
        data['branch_code'] = (
            bc_match.group(1) if bc_match else ""
        )
        if bc_match:
            after = bank_full[bc_match.end():].strip().lstrip(', ')
            branch_parts = re.split(r'[.,]', after)
            data['branch_name'] = (
                branch_parts[0].strip() if branch_parts else ""
            )
        else:
            data['branch_name'] = ""
    else:
        data['bank_name'] = ""
        data['branch_code'] = ""
        data['branch_name'] = ""

    data['account_type'] = ""

    gpf_acc_match = re.search(
        r'GPF\s+A/C\s+No:\s*(.+?)\s+Interest', text
    )
    if gpf_acc_match:
        val = gpf_acc_match.group(1).strip()
        data['gpf_account_no'] = val if val else ""
    else:
        data['gpf_account_no'] = ""

    data['ppo_no'] = ""

    ntn_match = re.search(r'NTN:\s*(\S+)', text)
    if ntn_match:
        ntn_val = ntn_match.group(1).strip()
        if len(ntn_val) > 1 and ntn_val not in ['Date']:
            data['ntn_no'] = ntn_val
        else:
            data['ntn_no'] = ""
    else:
        data['ntn_no'] = ""

    entry_match = re.search(
        r'Entry\s+into\s+Govt\.\s+Service:\s*([\d.]+)', text
    )
    data['date_of_appointment'] = (
        format_date(entry_match.group(1)) if entry_match else ""
    )
    data['date_of_entry'] = data['date_of_appointment']
    data['date_of_retirement'] = ""
    data['retirement_order_no'] = ""
    data['retirement_order_date'] = ""
    data['lwp_days'] = ""
    data['lpr_days'] = ""
    data['leave_taken_days'] = ""
    data['date_of_regularization'] = ""
    data['date_of_death'] = ""

    # ===== ALLOWANCES (mapped via PAY_ALLOWANCE_MAP, aligned with app codes) =====
    # Initialize all mapped allowance fields to empty
    for field in set(PAY_ALLOWANCE_MAP.values()):
        data[field] = ""

    # Apply canonical mappings
    for code, amount in pay_wages.items():
        if code in PAY_ALLOWANCE_MAP:
            field = PAY_ALLOWANCE_MAP[code]
            data[field] = amount

    # Legacy / alternate codes handling (PDF variants not in PAY_ALLOWANCE_MAP)
    # Medical Allowance sometimes appears as 1947 instead of 1300
    if not data.get('ma'):
        if '1947' in pay_wages:
            data['ma'] = pay_wages['1947']
        elif '1300' in pay_wages:
            data['ma'] = pay_wages['1300']

    # UAA sometimes appears under 1924/1925; 1923 is canonical
    if not data.get('uaa'):
        if '1925' in pay_wages:
            data['uaa'] = pay_wages['1925']
        elif '1924' in pay_wages:
            data['uaa'] = pay_wages['1924']
        elif '1923' in pay_wages:
            data['uaa'] = pay_wages['1923']

    # Special allowances that are not part of PAY_ALLOWANCE_MAP but exist in some payrolls
    if '2315' in pay_wages and not data.get('spl_allow'):
        data['spl_allow'] = pay_wages['2315']
    if '1550' in pay_wages and not data.get('spl_allow_female'):
        data['spl_allow_female'] = pay_wages['1550']
    if '1238' in pay_wages and not data.get('charge_allow'):
        data['charge_allow'] = pay_wages['1238']

    # Derived / combined Adhoc Reliefs according to app mapping
    if not data.get('adhoc_2023_35'):
        data['adhoc_2023_35'] = pay_wages.get('2379', pay_wages.get('2378', ""))
    if not data.get('adhoc_2024_25'):
        data['adhoc_2024_25'] = pay_wages.get('2394', pay_wages.get('2393', ""))

    # "Other" = all pay codes not explicitly mapped
    mapped_pay_codes = set(PAY_ALLOWANCE_MAP.keys()) | {
        '1947', '1924', '1925', '2315', '1550', '1238', '2379', '2394'
    }
    other_total = 0
    for code, amount in pay_wages.items():
        if code not in mapped_pay_codes:
            other_total += amount
    data['other'] = other_total if other_total > 0 else ""

    # Basic pay / last_basic_pay / last_pay_with_increment consistency
    basic = data.get('basic_pay') or pay_wages.get('0001', "")
    data['basic_pay'] = basic
    data['last_basic_pay'] = basic
    data['last_pay_with_increment'] = basic

    # ===== DEDUCTIONS (mapped via DEDUCTION_MAP) =====
    # Initialize mapped deduction fields
    for field in set(DEDUCTION_MAP.values()):
        if field not in data:
            data[field] = ""

    # Apply canonical deduction mappings
    for code, amount in deduct_wages.items():
        if code in DEDUCTION_MAP:
            field = DEDUCTION_MAP[code]
            # Multiple GPF subscription codes all map to gpf_sub; accumulate largest value
            if field == 'gpf_sub':
                existing = data.get('gpf_sub') or 0
                data['gpf_sub'] = max(existing, amount)
            else:
                data[field] = amount

    # Legacy / alternate GPF subscription codes (older formats)
    if not data.get('gpf_sub'):
        for gc in ['3017', '3016', '3015', '3014', '3006', '3003']:
            if gc in deduct_wages:
                data['gpf_sub'] = deduct_wages[gc]
                break

    # Map canonical GPF (total) from gpf_sub when available
    data['gpf'] = data.get('gpf') or data.get('gpf_sub') or ""

    # Other fixed deductions
    data['gpf_advance'] = data.get('gpf_advance', "")
    data['bf'] = data.get('bf', "")
    data['eef'] = data.get('eef', "")

    # RB Death: also consider legacy code 3534
    if not data.get('rb_death'):
        if '3534' in deduct_wages:
            data['rb_death'] = deduct_wages['3534']

    data['adl_g_insurance'] = data.get('adl_g_insurance', "")
    data['group_insurance'] = data.get('group_insurance', "")
    # Income tax (earning-side mirror) – we keep income_tax from deductions
    data['income_tax'] = deduct_wages.get('3609', data.get('income_tax', ""))
    data['recovery'] = deduct_wages.get('3621', data.get('recovery', ""))
    data['edu_rop'] = data.get('edu_rop', "")

    # ===== LOANS =====
    hba_match = re.search(
        r'6501\s+.+?([\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})\s+([\d,]+\.\d{2})',
        loan_section
    )
    data['hba_loan_instal'] = (
        clean_amount(hba_match.group(2)) if hba_match else ""
    )

    gpf_loan_match = re.search(
        r'6505\s+.+?([\d,]+\.\d{2})\s+(-?[\d,]+\.\d{2})\s+([\d,]+\.\d{2})',
        loan_section
    )
    data['gpf_loan_instal'] = (
        clean_amount(gpf_loan_match.group(2))
        if gpf_loan_match else ""
    )

    # Additional import fields (leave blank or mirror where applicable, without overriding mapped values)
    if not data.get('gpf_sub') and data.get('gpf'):
        data['gpf_sub'] = data['gpf']
    if not data.get('income_tax_ded') and data.get('income_tax'):
        data['income_tax_ded'] = data['income_tax']

    for key in [
        'computer_allow', 'mphil_allow', 'entertainment_allow',
        'science_teaching_allow', 'weather_allow', 'special_allow_non_teaching',
        'adhoc_2015', 'adhoc_2016', 'adhoc_2022',
        'allowances_extra_json', 'deductions_extra_json', 'arrears_json',
        'commutation_portion', 'retirement_date_source',
        'beneficiary_name', 'beneficiary_relation', 'beneficiary_age', 'beneficiary_cnic',
        'beneficiary_bank_name', 'beneficiary_branch_name', 'beneficiary_account_no',
    ]:
        if key not in data:
            data[key] = ""

    for i in range(1, 7):
        data[f'family_{i}_name'] = ""
        data[f'family_{i}_relation'] = ""
        data[f'family_{i}_age'] = ""
        data[f'family_{i}_cnic'] = ""

    return data


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    pdf_files = [
        os.path.join(script_dir, f)
        for f in os.listdir(script_dir)
        if f.lower().endswith('.pdf')
    ]

    if not pdf_files:
        print("ERROR: No PDF files found in", script_dir)
        sys.exit(1)

    print(f"Found {len(pdf_files)} PDF file(s):")
    for pf in pdf_files:
        print(f"  - {os.path.basename(pf)}")

    # Columns aligned with app CSV import/export schema (CSV_HEADERS)
    columns = [
        # Identity
        'name', 'designation', 'bps', 'school_full_name', 'office_name', 'staff_type', 'status',
        'cnic_no', 'personal_no', 'mobile_no', 'father_name',
        'dob', 'nationality', 'address', 'district', 'tehsil', 'ddo_code',
        'bank_ac_no', 'bank_name', 'branch_name', 'branch_code', 'account_type', 'gpf_account_no', 'ppo_no',
        'ntn_no', 'employment_category', 'designation_full',
        'gender',

        # Service
        'date_of_appointment', 'date_of_entry', 'date_of_retirement',
        'retirement_order_no', 'retirement_order_date',
        'lwp_days', 'lpr_days', 'leave_taken_days',
        'date_of_regularization', 'date_of_death',

        # Financials - Allowances
        'basic_pay', 'last_basic_pay', 'last_pay_with_increment', 'p_pay', 'hra', 'ca', 'ma', 'uaa',
        'spl_allow', 'teaching_allow', 'spl_allow_female', 'spl_allow_disable',
        'integrated_allow', 'charge_allow', 'wa', 'dress_allow', 'other',
        'computer_allow', 'mphil_allow', 'entertainment_allow', 'science_teaching_allow', 'weather_allow', 'special_allow_non_teaching',

        # Financials - Adhoc Reliefs
        'adhoc_2013', 'adhoc_2015', 'adhoc_2016', 'adhoc_2022', 'adhoc_10pct', 'dra_2022kp', 'adhoc_2022_ps17',
        'adhoc_2023_35', 'adhoc_2024_25', 'adhoc_2025_10', 'dra_2025_15',

        # Financials - Deductions & Loans
        'gpf', 'gpf_sub', 'gpf_advance', 'bf', 'eef', 'rb_death', 'adl_g_insurance', 'group_insurance', 'income_tax', 'income_tax_ded',
        'recovery', 'edu_rop', 'hba_loan_instal', 'gpf_loan_instal',
        'allowances_extra_json', 'deductions_extra_json', 'arrears_json',

        # Family Members (1-6)
        'family_1_name', 'family_1_relation', 'family_1_age', 'family_1_cnic',
        'family_2_name', 'family_2_relation', 'family_2_age', 'family_2_cnic',
        'family_3_name', 'family_3_relation', 'family_3_age', 'family_3_cnic',
        'family_4_name', 'family_4_relation', 'family_4_age', 'family_4_cnic',
        'family_5_name', 'family_5_relation', 'family_5_age', 'family_5_cnic',
        'family_6_name', 'family_6_relation', 'family_6_age', 'family_6_cnic',

        # Extras & Beneficiary
        'commutation_portion', 'retirement_date_source',
        'beneficiary_name', 'beneficiary_relation', 'beneficiary_age', 'beneficiary_cnic',
        'beneficiary_bank_name', 'beneficiary_branch_name', 'beneficiary_account_no',
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Data"

    header_font = Font(
        name='Calibri', bold=True, size=11, color='FFFFFF'
    )
    header_fill = PatternFill(
        start_color='4472C4', end_color='4472C4', fill_type='solid'
    )
    header_align = Alignment(
        horizontal='center', vertical='center', wrap_text=True
    )
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    data_font = Font(name='Calibri', size=10)
    data_align = Alignment(vertical='center')

    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.freeze_panes = 'A2'

    row_num = 2
    total = 0

    for pdf_file in pdf_files:
        print(f"\nProcessing: {os.path.basename(pdf_file)}")
        try:
            reader = PdfReader(pdf_file)
            print(f"  Pages: {len(reader.pages)}")

            for page_idx in range(len(reader.pages)):
                page_text = reader.pages[page_idx].extract_text()
                if not page_text:
                    continue
                if 'Personal Information' not in page_text:
                    continue

                try:
                    emp = extract_employee_data(page_text)

                    # Skip ghost/incomplete records (no name, no personnel number, or missing BPS)
                    if not emp.get('name') or not emp.get('personal_no') or not emp.get('bps'):
                        print(f"  Page {page_idx+1}: SKIP (incomplete/ghost record)")
                        continue

                    # Verify by summing all allowances
                    allow_fields = [
                        'last_basic_pay', 'p_pay', 'hra', 'ca',
                        'ma', 'uaa', 'spl_allow', 'teaching_allow',
                        'spl_allow_female', 'spl_allow_disable',
                        'integrated_allow', 'charge_allow', 'wa',
                        'dress_allow', 'other',
                        'adhoc_2013', 'adhoc_10pct', 'dra_2022kp',
                        'adhoc_2022_ps17', 'adhoc_2023_35',
                        'adhoc_2024_25', 'adhoc_2025_10',
                        'dra_2025_15',
                    ]
                    calc_gross = 0
                    for af in allow_fields:
                        v = emp.get(af, "")
                        if isinstance(v, (int, float)):
                            calc_gross += v

                    # Get actual gross from PDF
                    gross_match = re.search(
                        r'Gross\s+Pay\s+\(Rs\.\):\s+([\d,]+\.\d{2})',
                        page_text
                    )
                    pdf_gross = 0
                    if gross_match:
                        pdf_gross = float(
                            gross_match.group(1).replace(',', '')
                        )

                    for col_idx, col_name in enumerate(columns, 1):
                        val = emp.get(col_name, "")
                        cell = ws.cell(
                            row=row_num, column=col_idx, value=val
                        )
                        cell.font = data_font
                        cell.alignment = data_align
                        cell.border = thin_border

                    diff = pdf_gross - calc_gross
                    status_mark = "OK" if abs(diff) < 1 else f"DIFF={diff}"

                    print(
                        f"  Page {page_idx+1}: "
                        f"{emp.get('name','?')} | "
                        f"{emp.get('designation','')} | "
                        f"BPS-{emp.get('bps','')} | "
                        f"Basic:{emp.get('last_basic_pay','')} | "
                        f"CA:{emp.get('ca','')} | "
                        f"Adhoc13:{emp.get('adhoc_2013','')} | "
                        f"Gross:PDF={pdf_gross},Calc={calc_gross} "
                        f"[{status_mark}]"
                    )
                    row_num += 1
                    total += 1
                except Exception as e:
                    print(f"  Page {page_idx+1}: ERROR - {e}")
        except Exception as e:
            print(f"  Error: {e}")

    for col_idx in range(1, len(columns) + 1):
        cl = ws.cell(row=1, column=col_idx).column_letter
        mx = len(
            str(ws.cell(row=1, column=col_idx).value or "")
        ) + 2
        for r in range(2, row_num):
            v = ws.cell(row=r, column=col_idx).value
            if v:
                mx = max(mx, len(str(v)) + 2)
        ws.column_dimensions[cl].width = min(mx, 30)

    output = os.path.join(script_dir, "payroll_extracted.xlsx")
    wb.save(output)

    print(f"\n{'='*55}")
    print(f"  DONE! Employees: {total}")
    print(f"  File: {output}")
    print(f"{'='*55}")


if __name__ == "__main__":
    main()
